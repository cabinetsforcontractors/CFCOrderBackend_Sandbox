"""
supplier_orders.py
The SPINE of the auto-ordering lane: one row per order x warehouse walking
through the supplier-order lifecycle, plus the dispatch engine that generates
and (guardedly) sends each supplier's order artifact.

Status flow:
    pending -> sent | prepared | blocked
    sent/prepared -> confirmed | discrepancy      (verification step)
    confirmed -> scheduled -> picked_up -> delivered -> invoice_verified

SUPPLIER EMAIL FORMAT (William 2026-07-17, supersedes the LI/GHI customer-info
exception for EMAIL BODIES — the GHI xlsx sheet keeps its Ship To field):
  - NO customer info in any supplier email (no name/address/phone)
  - NO "Our SKU" column — supplier sees THEIR SKUs only
  - our door-style names STRIPPED from descriptions ("TRUE WHITE SHAKER,
    WALL FILLER 3''W" -> "WALL FILLER 3''W")
  - wording: "Please process our order PO {id}: ({their door name}, {their
    presku})" — door info from SUPPLIER_DOOR_INFO (unknown lines omit it)
  - footer: "Total Qty All SKUS: ##"
  - TEMPLATE v2 (William's own GHI send, 2026-07-17): greet the supplier
    CONTACT by first name ("Hey Kathryn,"), William's signature block.
    Attachment emails say "See attached for our PO {id}: (door, presku)."
    v2.1: close with an ACTIVE stock ask — "Please check for any
    out-of-stock items and let us know." (William: nudge them to check,
    don't just ask a yes/no question.)

Channels (SUPPLIER_ORDER_CHANNELS_20260716.md + William rulings):
  EMAIL-AUTO: GHI (xlsx sheet), LI, Cabinet & Stone, DuraStone, Love-Milestone
  PORTAL-PREPARED (emailed TO US for the ~2-min upload): ROC (quick-order
  CSV), DL, L&C Cabinetry, Linda, Go Bravura

ROC QUICK-ORDER DIALECT: store-prefixed SKUs (SNW-TK8); ROC_STORE_PREFIX maps
our line -> their store line; unknown lines BLOCK. COMPANION RULE: easy-reach
bases auto-add the free lazy-susan tray A-BER-B. QUANTITY PARITY: CSV units
must equal website-order units or the warehouse blocks.

TEMPLATE-IN-DB (2026-07-17): the GHI order-sheet template lives in the
supplier_templates table (upload via POST /supplier-orders/template/GHI), so
dispatch builds GHI sheets with no env path and no manual step.
GHI_TEMPLATE_PATH env still wins when set. normalize_ghi_template() repairs
GHI's annotated cells (JKFTS-Jiffy Kit-...) before the sheet filler runs.

PAYMENT TRIGGER (payment_triggers.py): auto-dispatch only when
AUTO_DISPATCH_ENABLED=true AND payment matches order total within $1.

All sends go through the Gmail path with the EMAIL_ALLOWLIST guard —
in the beta everything redirects to the test inbox.
"""

import base64
import io
import json
import os
import re
from datetime import datetime, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import Dict, List, Optional

from config import SUPPLIER_INFO
from db_helpers import get_db

INTERNAL_ALERT_EMAIL = os.environ.get("WAREHOUSE_NOTIFICATION_EMAIL",
                                      "cabinetsforcontractors@gmail.com").strip()

STATUSES = ("pending", "sent", "prepared", "blocked", "confirmed", "discrepancy",
            "scheduled", "picked_up", "delivered", "invoice_verified", "canceled")

# warehouse key (as used by /freight/supplier-sheet) -> channel config
SUPPLIER_CHANNELS = {
    "GHI":             {"mode": "email_auto", "artifact": "ghi_xlsx"},
    "LI":              {"mode": "email_auto", "artifact": "po_email"},
    "Cabinet & Stone": {"mode": "email_auto", "artifact": "po_email"},
    "DuraStone":       {"mode": "email_auto", "artifact": "po_email"},
    "Love-Milestone":  {"mode": "email_auto", "artifact": "po_email"},
    "ROC":             {"mode": "portal_prepared", "artifact": "roc_csv"},
    "DL":              {"mode": "portal_prepared", "artifact": "po_email"},
    "L&C Cabinetry":   {"mode": "portal_prepared", "artifact": "po_email"},
    "Linda":           {"mode": "portal_prepared", "artifact": "po_email"},
    "Go Bravura":      {"mode": "portal_prepared", "artifact": "po_email"},
}

# our website line prefix -> ROC store line prefix (their quick-order SKUs).
# LNS -> SNW proven by William's portal upload test + confirmation #000040179.
ROC_STORE_PREFIX = {
    "LNS": "SNW",
}

# Easy-reach bases need the free lazy-susan TRAYS on ROC orders (William
# 2026-07-17): tray SKU A-BER-B, $0, same SKU for 33"/36", any color.
ROC_TRAY_SKU = "A-BER-B"
_ROC_EASY_REACH = re.compile(r"^B?ER(33|36)$")

# THE SUPPLIER'S door name + presku per (supplier, our line prefix) — used in
# supplier-facing emails ("their door name, their door presku", William
# 2026-07-17). Fill in as lines are ruled; unknown lines omit the door header.
SUPPLIER_DOOR_INFO = {
    ("DuraStone", "NSN"):      {"door_name": "Natural Wood",     "presku": "NW"},
    ("ROC", "LNS"):            {"door_name": "Natural Shaker",   "presku": "SNW"},
    ("GHI", "AKS"):            {"door_name": "Frontier",         "presku": "FTS"},
    ("GHI", "APW"):            {"door_name": "Rustic Walnut",    "presku": "RWS"},
    ("GHI", "GRSH"):           {"door_name": "Stone Harbor Gray", "presku": "SHG"},
    ("GHI", "NOR"):            {"door_name": "Nantucket",        "presku": "NTL"},
    ("GHI", "SNS"):            {"door_name": "Sonona Sand",      "presku": "SNS"},
    ("GHI", "SNW"):            {"door_name": "Sonona Wheat",     "presku": "SNW"},
    ("Love-Milestone", "SB"):  {"door_name": "Sage Breeze",      "presku": "SB"},
}

# William's signature block for supplier-facing emails (his own GHI send,
# 2026-07-17, is the template)
SIGNATURE_HTML = (
    "<p>--<br>William Prince<br>Cabinets For Contractors<br>"
    "<a href='https://www.CabinetsForContractors.net'>www.CabinetsForContractors.net</a><br>"
    "(770) 990-4885</p>")

# v2.1 closing line (William 2026-07-17): actively ask the supplier to CHECK
# stock — not a yes/no question
STOCK_CHECK_ASK = "<p>Please check for any out-of-stock items and let us know.</p>"


def supplier_greeting(warehouse: str) -> str:
    """'Hey Kathryn,' from SUPPLIER_INFO contact's first name; 'Hello,' when
    no contact is on file."""
    contact = (SUPPLIER_INFO.get(warehouse) or {}).get("contact", "") or ""
    first = contact.strip().split(" ")[0].strip(",")
    if first and first.isalpha():
        return f"Hey {first},"
    return "Hello,"


def strip_our_door_name(product_name: str) -> str:
    """'TRUE WHITE SHAKER, WALL FILLER 3''W X 36''H' -> 'WALL FILLER 3''W X 36''H'
    (our door-style name must not appear in supplier correspondence)."""
    s = (product_name or "").strip()
    if "," in s:
        head, _, rest = s.partition(",")
        # only strip when the head looks like a door-style name (letters/spaces)
        if rest.strip() and not any(ch.isdigit() for ch in head):
            return rest.strip()
    return s


def door_info_for(supplier: str, items: List[Dict]) -> Optional[Dict]:
    """The supplier's door name/presku for this warehouse's lines (first known)."""
    for i in items:
        pre = (i.get("website_sku") or "").split("-")[0].upper()
        info = SUPPLIER_DOOR_INFO.get((supplier, pre))
        if info:
            return info
    return None


# =============================================================================
# TABLE
# =============================================================================

def ensure_supplier_orders_table(conn):
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS supplier_orders (
                id SERIAL PRIMARY KEY,
                order_id VARCHAR(20) NOT NULL,
                warehouse VARCHAR(100) NOT NULL,
                status VARCHAR(30) DEFAULT 'pending',
                mode VARCHAR(30),
                artifact_type VARCHAR(30),
                artifact_filename VARCHAR(200),
                line_count INTEGER,
                untranslated_count INTEGER DEFAULT 0,
                sent_to VARCHAR(200),
                sent_at TIMESTAMP WITH TIME ZONE,
                confirmed_at TIMESTAMP WITH TIME ZONE,
                supplier_doc_ref VARCHAR(100),
                discrepancy_json TEXT,
                note TEXT,
                created_at TIMESTAMP WITH TIME ZONE DEFAULT NOW(),
                updated_at TIMESTAMP WITH TIME ZONE DEFAULT NOW(),
                UNIQUE (order_id, warehouse)
            )
        """)
        cur.execute("ALTER TABLE supplier_orders ADD COLUMN IF NOT EXISTS revision_requested_at TIMESTAMP WITH TIME ZONE")
        cur.execute("ALTER TABLE supplier_orders ADD COLUMN IF NOT EXISTS followup_sent_at TIMESTAMP WITH TIME ZONE")
        cur.execute("ALTER TABLE supplier_orders ADD COLUMN IF NOT EXISTS no_response_alerted_at TIMESTAMP WITH TIME ZONE")
        conn.commit()


# =============================================================================
# SUPPLIER TEMPLATE STORE (template-in-DB, 2026-07-17)
# =============================================================================

def ensure_supplier_templates_table(conn):
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS supplier_templates (
                supplier VARCHAR(50) PRIMARY KEY,
                filename VARCHAR(200),
                content BYTEA NOT NULL,
                uploaded_at TIMESTAMP WITH TIME ZONE DEFAULT NOW()
            )
        """)
        conn.commit()


def save_supplier_template(supplier: str, filename: str, content: bytes) -> Dict:
    import psycopg2
    with get_db() as conn:
        ensure_supplier_templates_table(conn)
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO supplier_templates (supplier, filename, content, uploaded_at)
                VALUES (%s, %s, %s, NOW())
                ON CONFLICT (supplier) DO UPDATE SET
                    filename = EXCLUDED.filename,
                    content = EXCLUDED.content,
                    uploaded_at = NOW()
            """, (supplier.upper(), filename, psycopg2.Binary(content)))
            conn.commit()
    return {"supplier": supplier.upper(), "filename": filename,
            "bytes": len(content)}


def get_supplier_template(supplier: str):
    """-> (filename, content_bytes) or None."""
    with get_db() as conn:
        ensure_supplier_templates_table(conn)
        with conn.cursor() as cur:
            cur.execute("""SELECT filename, content FROM supplier_templates
                           WHERE supplier = %s""", (supplier.upper(),))
            row = cur.fetchone()
            if not row:
                return None
            return row[0], bytes(row[1])


def list_supplier_templates() -> List[Dict]:
    from psycopg2.extras import RealDictCursor
    with get_db() as conn:
        ensure_supplier_templates_table(conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""SELECT supplier, filename, octet_length(content) AS bytes,
                                  uploaded_at
                           FROM supplier_templates ORDER BY supplier""")
            return cur.fetchall()


def normalize_ghi_template(tpl_bytes: bytes) -> bytes:
    """Repair GHI's annotated cells BEFORE the sheet filler scans them:
    'JKFTS-Jiffy Kit-FTS-(M883-3369)' -> 'JKFTS'. The filler's cell matcher
    requires labels to END with the tab code, so annotated rows (the jiffy
    kits) were never placeable (order 5693 finding, corrections queue
    2026-07-17). Only rewrites cabinet-column cells whose compact text is
    token+tab+'-annotation'."""
    import openpyxl
    try:
        from supplier_doc_parser import GHI_LINE_TAB
        tabs = {t.upper() for t in GHI_LINE_TAB.values()}
    except Exception:
        tabs = {"FTS", "RWS", "SHG", "NTL", "SNS", "SNW"}
    wb = openpyxl.load_workbook(io.BytesIO(tpl_bytes))
    changed = 0
    for name in wb.sheetnames:
        tab = name.strip().upper()
        if tab not in tabs:
            continue
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200)):
            for c in row:
                if not isinstance(c.value, str) or c.column <= 1:
                    continue
                compact = re.sub(r"\s+", "", c.value).upper()
                m = re.match(rf"^([A-Z0-9/.]+{tab})-", compact)
                if m and not compact.endswith(tab):
                    c.value = m.group(1)
                    changed += 1
    if not changed:
        return tpl_bytes
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =============================================================================
# GUARDED MAILER (allowlist + optional attachment + CC)
# =============================================================================

def _send_email(order_id: str, to_email: str, subject: str, html: str,
                triggered_by: str, attachment: Optional[Dict] = None,
                cc: Optional[str] = None) -> Dict:
    """Guarded Gmail send. attachment = {'filename','content'(bytes),'mime'}."""
    from config import GMAIL_SEND_ENABLED
    from email_sender import _log_email_event
    from gmail_sync import get_gmail_access_token

    if not to_email or "@" not in to_email:
        return {"success": False, "error": f"invalid email: {to_email}"}
    if not GMAIL_SEND_ENABLED:
        return {"success": False, "error": "GMAIL_SEND_ENABLED=false", "dry_run": True}
    allowlist = os.environ.get("EMAIL_ALLOWLIST", "").strip()
    if allowlist:
        allowed = {e.strip().lower() for e in allowlist.split(",") if e.strip()}
        if to_email.lower() not in allowed:
            redirect = os.environ.get("INTERNAL_SAFETY_EMAIL", "").strip()
            if redirect:
                print(f"[DISPATCH-GUARD] redirected {to_email} -> {redirect} order={order_id}")
                to_email = redirect
            else:
                print(f"[DISPATCH-GUARD] blocked {to_email} order={order_id}")
                return {"success": False, "error": "recipient not in EMAIL_ALLOWLIST",
                        "dry_run": True, "original_to": to_email}
        if cc and cc.lower() not in allowed:
            cc = os.environ.get("INTERNAL_SAFETY_EMAIL", "").strip() or None
            if cc and cc.lower() == to_email.lower():
                cc = None  # avoid To == Cc after double redirect
    try:
        token = get_gmail_access_token()
        if not token:
            return {"success": False, "error": "no Gmail access token"}
        msg = MIMEMultipart("mixed")
        msg["To"] = to_email
        if cc:
            msg["Cc"] = cc
        msg["From"] = "William Prince — Cabinets For Contractors <william@cabinetsforcontractors.net>"
        msg["Subject"] = subject
        msg.attach(MIMEText(html, "html"))
        if attachment and attachment.get("content"):
            main, _, sub = (attachment.get("mime") or "application/octet-stream").partition("/")
            part = MIMEBase(main, sub or "octet-stream")
            part.set_payload(attachment["content"])
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", "attachment",
                            filename=attachment.get("filename", "attachment"))
            msg.attach(part)
        import urllib.request
        req = urllib.request.Request(
            "https://gmail.googleapis.com/gmail/v1/users/me/messages/send",
            data=json.dumps({"raw": base64.urlsafe_b64encode(msg.as_bytes()).decode()}).encode(),
            method="POST")
        req.add_header("Authorization", f"Bearer {token}")
        req.add_header("Content-Type", "application/json")
        with urllib.request.urlopen(req, timeout=60) as resp:
            message_id = json.loads(resp.read().decode()).get("id")
        _log_email_event(order_id=order_id, template_id="supplier_dispatch",
                         to_email=to_email, subject=subject, message_id=message_id,
                         triggered_by=triggered_by, source="email_send")
        return {"success": bool(message_id), "message_id": message_id,
                "to": to_email, "cc": cc}
    except Exception as e:
        return {"success": False, "error": str(e)}


# =============================================================================
# ARTIFACT BUILDERS
# =============================================================================

def _supplier_lines_table(items: List[Dict]) -> str:
    """Supplier-facing table: THEIR SKU + qty + sanitized description only —
    no 'Our SKU' column, our door names stripped (William 2026-07-17)."""
    rows = "".join(
        f"<tr><td style='padding:4px 10px;border-bottom:1px solid #ddd;'>{i['quantity']}</td>"
        f"<td style='padding:4px 10px;border-bottom:1px solid #ddd;'><strong>{i['supplier_sku']}</strong></td>"
        f"<td style='padding:4px 10px;border-bottom:1px solid #ddd;'>{strip_our_door_name(i.get('product_name', ''))}</td></tr>"
        for i in items)
    return (f"<table style='border-collapse:collapse;font-size:13px;'>"
            f"<tr style='background:#f2f2f2;'><th align='left' style='padding:4px 10px;'>Qty</th>"
            f"<th align='left' style='padding:4px 10px;'>SKU</th>"
            f"<th align='left' style='padding:4px 10px;'>Description</th></tr>{rows}</table>")


def _internal_lines_table(items: List[Dict]) -> str:
    """Internal table (upload-needed emails to US): includes Our SKU."""
    rows = "".join(
        f"<tr><td style='padding:4px 10px;border-bottom:1px solid #ddd;'>{i['quantity']}</td>"
        f"<td style='padding:4px 10px;border-bottom:1px solid #ddd;'><strong>{i['supplier_sku']}</strong></td>"
        f"<td style='padding:4px 10px;border-bottom:1px solid #ddd;color:#888;'>{i['website_sku']}</td>"
        f"<td style='padding:4px 10px;border-bottom:1px solid #ddd;'>{i.get('product_name', '')}</td></tr>"
        for i in items)
    return (f"<table style='border-collapse:collapse;font-size:13px;'>"
            f"<tr style='background:#f2f2f2;'><th align='left' style='padding:4px 10px;'>Qty</th>"
            f"<th align='left' style='padding:4px 10px;'>SKU</th>"
            f"<th align='left' style='padding:4px 10px;'>Our SKU</th>"
            f"<th align='left' style='padding:4px 10px;'>Description</th></tr>{rows}</table>")


def build_po_email(order_id: str, warehouse: str, wdata: Dict) -> Dict:
    """Supplier-facing PO email: no customer info, no Our SKU column, our
    door names stripped, their door name/presku in the opening line,
    'Total Qty All SKUS' footer (William 2026-07-17). Template v2.1: contact
    first-name greeting, active stock-check ask, William's signature."""
    items = wdata["items"]
    total_units = sum(int(i.get("quantity") or 0) for i in items)
    door = door_info_for(warehouse, items)
    door_txt = f" ({door['door_name']}, {door['presku']})" if door else ""
    html = (f"<div style='font-family:Arial,sans-serif;font-size:14px;'>"
            f"<p>{supplier_greeting(warehouse)}</p>"
            f"<p>Please process our order <strong>PO {order_id}</strong>:{door_txt}</p>"
            f"{_supplier_lines_table(items)}"
            f"<p><strong>Total Qty All SKUS: {total_units}</strong></p>"
            f"{STOCK_CHECK_ASK}"
            f"{SIGNATURE_HTML}</div>")
    return {"html": html, "attachment": None, "units": total_units,
            "subject": f"PO {order_id} - Cabinets For Contractors"}


def build_roc_csv(order_id: str, wdata: Dict) -> Dict:
    """ROC quick-order CSV (store-prefixed SKUs) — INTERNAL email to us with
    the upload instructions. Parity gate + free lazy-susan tray companions."""
    rows = []
    unknown = []
    order_units = 0
    csv_units = 0
    tray_qty = 0
    for i in wdata["items"]:
        qty = int(float(i.get("quantity") or 0))
        order_units += qty
        token = (i["supplier_sku"] or "").strip()
        our_prefix = (i["website_sku"] or "").split("-")[0].upper()
        if "-" in token and token.split("-")[0].upper() in ROC_STORE_PREFIX.values():
            store_sku = token  # map already carries a store-prefixed SKU
        else:
            store_prefix = ROC_STORE_PREFIX.get(our_prefix)
            if not store_prefix:
                unknown.append(i["website_sku"])
                continue
            store_sku = f"{store_prefix}-{token}"
        rows.append(f"{store_sku},{qty}\n")
        csv_units += qty
        if _ROC_EASY_REACH.match(token.upper()):
            tray_qty += qty
    if unknown:
        return {"error": (f"ROC store prefix unknown for line(s) of: "
                          f"{', '.join(unknown[:10])} — add to ROC_STORE_PREFIX "
                          f"after confirming on their portal")}
    if csv_units != order_units:
        return {"error": (f"QUANTITY PARITY FAILED: website order has "
                          f"{order_units} units for ROC but the CSV built "
                          f"{csv_units} — refusing to send")}
    if tray_qty:
        rows.append(f"{ROC_TRAY_SKU},{tray_qty}\n")
    csv_text = "sku,qty\n" + "".join(rows)
    tray_note = (f" Plus {tray_qty} x {ROC_TRAY_SKU} free lazy-susan trays "
                 f"auto-added (easy-reach rule)." if tray_qty else "")
    html = (f"<div style='font-family:Arial,sans-serif;font-size:14px;'>"
            f"<p><strong>UPLOAD NEEDED - ROC quick-order CSV for PO {order_id}</strong></p>"
            f"<p><strong>Quantity check:</strong> website order units = {order_units}, "
            f"CSV units = {csv_units} &#10003;{tray_note}</p>"
            f"<p><strong>Total Qty All SKUS: {order_units}</strong></p>"
            f"<p>Attached: the quick-order file. Upload at "
            f"roccabinetry.com/quick-order, ENTER PO {order_id} in their "
            f"PO/reference field, then mark this supplier order as sent. "
            f"AFTER upload: copy the whole cart page and POST it to "
            f"/supplier-orders/roc-stock-paste to catch out-of-stock flags "
            f"(the CSV export does not carry them).</p>"
            f"{_internal_lines_table(wdata['items'])}</div>")
    return {"html": html, "units": order_units, "companions": tray_qty,
            "attachment": {"filename": f"ROC_order_{order_id}.csv",
                           "content": csv_text.encode(), "mime": "text/csv"},
            "subject": f"UPLOAD NEEDED: ROC quick-order CSV - PO {order_id}"}


def build_ghi_xlsx(order_id: str, wdata: Dict) -> Dict:
    """GHI order sheet (the 5707.xlsx format). Template resolution:
    GHI_TEMPLATE_PATH env (wins when set) -> supplier_templates DB row
    (POST /supplier-orders/template/GHI) -> error. The email body is
    supplier-facing (no customer info); the sheet itself keeps its Ship To
    field (William-approved artifact)."""
    tpl = None
    tpl_path = os.environ.get("GHI_TEMPLATE_PATH", "").strip()
    if tpl_path and os.path.exists(tpl_path):
        with open(tpl_path, "rb") as f:
            tpl = f.read()
    if tpl is None:
        stored = get_supplier_template("GHI")
        if stored:
            tpl = stored[1]
    if tpl is None:
        return {"error": "no GHI template on this environment — upload one via "
                         "POST /supplier-orders/template/GHI (or set GHI_TEMPLATE_PATH)"}
    tpl = normalize_ghi_template(tpl)

    import supplier_doc_parser as sdp
    with get_db() as conn:
        fwd = sdp.build_forward_map(conn)
    items = [{"website_sku": i["website_sku"], "quantity": i["quantity"]}
             for i in wdata["items"]]
    from psycopg2.extras import RealDictCursor
    with get_db() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT company_name, customer_name FROM orders WHERE order_id = %s",
                        (order_id,))
            o = cur.fetchone() or {}
    ship_to = (o.get("company_name") or o.get("customer_name") or "")
    xlsx, report = sdp.make_ghi_sheets(items, tpl, order_id, fwd,
                                       ship_to=f"{ship_to} / PO {order_id}".strip(" /"))
    if report["unplaced"] or report["unmapped_prefix"]:
        return {"error": f"GHI sheet needs review before sending: "
                         f"unplaced={report['unplaced']} unmapped={report['unmapped_prefix']}"}
    total_units = sum(int(i.get("quantity") or 0) for i in wdata["items"])
    door = door_info_for("GHI", wdata["items"])
    door_txt = f" ({door['door_name']}, {door['presku']})" if door else ""
    html = (f"<div style='font-family:Arial,sans-serif;font-size:14px;'>"
            f"<p>{supplier_greeting('GHI')}</p>"
            f"<p>See attached for our <strong>PO {order_id}</strong>:{door_txt}</p>"
            f"<p><strong>Total Qty All SKUS: {total_units}</strong></p>"
            f"{STOCK_CHECK_ASK}"
            f"{SIGNATURE_HTML}</div>")
    return {"html": html, "units": total_units,
            "attachment": {"filename": f"CFC_PO_{order_id}_GHI.xlsx",
                           "content": xlsx,
                           "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
            "subject": f"PO {order_id} - Cabinets For Contractors order sheet"}


# =============================================================================
# DISPATCH ENGINE
# =============================================================================

def dispatch_order(order_id: str, auto_send: bool = True,
                   dry_run: bool = False, triggered_by: str = "manual") -> Dict:
    """Generate every warehouse's artifact for an order and (unless dry_run)
    send email-auto channels to the supplier and portal-prepared channels to
    us. Rows land in supplier_orders with the resulting status. Untranslated
    lines block that warehouse — DO NOT auto-send half an order."""
    from freight_routes import get_supplier_sheet
    sheet = get_supplier_sheet(order_id, True)
    if sheet.get("status") != "ok":
        return {"status": "error", "message": sheet.get("message", "supplier-sheet failed")}

    results = {"status": "ok", "order_id": order_id, "dry_run": dry_run,
               "warehouses": {}, "blocked": 0, "sent": 0, "prepared": 0}
    with get_db() as conn:
        ensure_supplier_orders_table(conn)

    for wh, wdata in (sheet.get("warehouses") or {}).items():
        ch = SUPPLIER_CHANNELS.get(wh, {"mode": "portal_prepared",
                                        "artifact": "po_email"})
        sinfo = SUPPLIER_INFO.get(wh, {})
        wres = {"mode": ch["mode"], "artifact": ch["artifact"],
                "lines": len(wdata["items"]),
                "untranslated": len(wdata["untranslated"])}

        if wh == "UNMAPPED" or wdata["untranslated"]:
            status = "blocked"
            note = (f"{len(wdata['untranslated'])} untranslated lines"
                    if wdata["untranslated"] else "unmapped warehouse")
            wres.update({"status": status, "note": note,
                         "untranslated_skus": [u["website_sku"]
                                               for u in wdata["untranslated"]][:20]})
            _upsert_row(order_id, wh, status, ch, wres, note=note)
            results["blocked"] += 1
            results["warehouses"][wh] = wres
            continue
        if not wdata["items"]:
            continue

        # build the artifact
        if ch["artifact"] == "ghi_xlsx":
            art = build_ghi_xlsx(order_id, wdata)
        elif ch["artifact"] == "roc_csv":
            art = build_roc_csv(order_id, wdata)
        else:
            art = build_po_email(order_id, wh, wdata)
        if art.get("error"):
            status = "blocked"
            wres.update({"status": status, "note": art["error"]})
            _upsert_row(order_id, wh, status, ch, wres, note=art["error"])
            results["blocked"] += 1
            results["warehouses"][wh] = wres
            continue

        wres["subject"] = art["subject"]
        wres["attachment"] = (art.get("attachment") or {}).get("filename")
        if art.get("units") is not None:
            wres["units"] = art["units"]
        if art.get("companions"):
            wres["companions"] = art["companions"]

        if dry_run:
            wres["status"] = "dry_run"
            wres["preview"] = art["html"][:1500]
            results["warehouses"][wh] = wres
            continue

        if ch["mode"] == "email_auto" and auto_send:
            to_addr = sinfo.get("email", "")
            send = _send_email(order_id, to_addr, art["subject"], art["html"],
                               triggered_by, art.get("attachment"))
            status = "sent" if send.get("success") else "blocked"
            note = None if send.get("success") else f"send failed: {send.get('error')}"
            wres.update({"status": status, "send": send, "sent_to": send.get("to")})
            results["sent" if status == "sent" else "blocked"] += 1
        else:
            # portal supplier (or auto_send off): prepared artifact comes TO US
            send = _send_email(order_id, INTERNAL_ALERT_EMAIL,
                               f"[ACTION] {art['subject']}" if ch["mode"] != "email_auto"
                               else f"[CONFIRM+SEND] {art['subject']}",
                               art["html"], triggered_by, art.get("attachment"))
            status = "prepared" if send.get("success") else "blocked"
            note = None if send.get("success") else f"send failed: {send.get('error')}"
            wres.update({"status": status, "send": send})
            results["prepared" if status == "prepared" else "blocked"] += 1

        _upsert_row(order_id, wh, status, ch, wres, note=note,
                    sent_to=wres.get("sent_to") or wres.get("send", {}).get("to"))
        results["warehouses"][wh] = wres

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO order_events (order_id, event_type, event_data, source)
                VALUES (%s, 'supplier_dispatch', %s, 'supplier_orders')
            """, (order_id, json.dumps({
                "dry_run": dry_run, "triggered_by": triggered_by,
                "summary": {k: results[k] for k in ("sent", "prepared", "blocked")},
                "warehouses": {w: r.get("status") for w, r in results["warehouses"].items()},
            })))
            conn.commit()
    return results


def _upsert_row(order_id: str, warehouse: str, status: str, ch: Dict,
                wres: Dict, note: str = None, sent_to: str = None):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO supplier_orders
                    (order_id, warehouse, status, mode, artifact_type,
                     artifact_filename, line_count, untranslated_count,
                     sent_to, sent_at, note, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s,
                        CASE WHEN %s IN ('sent', 'prepared') THEN NOW() END,
                        %s, NOW())
                ON CONFLICT (order_id, warehouse) DO UPDATE SET
                    status = EXCLUDED.status,
                    mode = EXCLUDED.mode,
                    artifact_type = EXCLUDED.artifact_type,
                    artifact_filename = EXCLUDED.artifact_filename,
                    line_count = EXCLUDED.line_count,
                    untranslated_count = EXCLUDED.untranslated_count,
                    sent_to = COALESCE(EXCLUDED.sent_to, supplier_orders.sent_to),
                    sent_at = COALESCE(EXCLUDED.sent_at, supplier_orders.sent_at),
                    note = EXCLUDED.note,
                    updated_at = NOW()
            """, (order_id, warehouse, status, ch["mode"], ch["artifact"],
                  wres.get("attachment"), wres.get("lines"), wres.get("untranslated", 0),
                  sent_to, status, note))
            conn.commit()


# =============================================================================
# PAYMENT TRIGGER GATE
# =============================================================================

def run_dispatch_on_payment(order_id: str, order_data: dict,
                            payment_amount: float) -> Dict:
    """Trigger 5 (William: 100% auto-send when payment is made) with two gates:
      - AUTO_DISPATCH_ENABLED=true (env; default OFF for the beta)
      - exact payment: |payment - order_total| <= $1.00. Fuzzy Gmail-matched
        payments must NEVER place supplier orders — those get a
        confirm-dispatch alert to William instead."""
    enabled = os.environ.get("AUTO_DISPATCH_ENABLED", "false").lower() == "true"
    order_total = float(order_data.get("order_total") or 0)
    exact = order_total > 0 and abs(float(payment_amount) - order_total) <= 1.00

    if enabled and exact:
        return dispatch_order(order_id, auto_send=True, dry_run=False,
                              triggered_by="payment_trigger")

    # gate closed -> create rows as pending + alert William
    reason = ("AUTO_DISPATCH_ENABLED=false" if not enabled else
              f"payment ${payment_amount:,.2f} does not exactly match order "
              f"total ${order_total:,.2f} (fuzzy match — human confirm required)")
    try:
        preview = dispatch_order(order_id, auto_send=False, dry_run=True,
                                 triggered_by="payment_trigger_gated")
    except Exception as e:
        preview = {"status": "error", "message": str(e)}
    _send_email(order_id, INTERNAL_ALERT_EMAIL,
                f"CONFIRM DISPATCH: order #{order_id} paid - supplier orders ready",
                f"<p>Payment received for order <strong>#{order_id}</strong> "
                f"(${payment_amount:,.2f}).</p>"
                f"<p><strong>Not auto-dispatched:</strong> {reason}</p>"
                f"<p>To send: POST /supplier-orders/dispatch/{order_id}</p>"
                f"<p>Warehouses: "
                f"{', '.join((preview.get('warehouses') or {}).keys()) or 'n/a'}</p>",
                triggered_by="payment_trigger_gated")
    return {"status": "gated", "reason": reason, "preview": preview}


# =============================================================================
# QUERIES / TRANSITIONS
# =============================================================================

def list_supplier_orders(order_id: str = None, status: str = None,
                         limit: int = 100) -> List[Dict]:
    from psycopg2.extras import RealDictCursor
    q = "SELECT * FROM supplier_orders WHERE TRUE"
    args = []
    if order_id:
        q += " AND order_id = %s"
        args.append(order_id)
    if status:
        q += " AND status = %s"
        args.append(status)
    q += " ORDER BY updated_at DESC LIMIT %s"
    args.append(min(int(limit), 500))
    with get_db() as conn:
        ensure_supplier_orders_table(conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(q, args)
            return cur.fetchall()


def set_status(row_id: int, status: str, note: str = None,
               supplier_doc_ref: str = None) -> Dict:
    if status not in STATUSES:
        return {"status": "error", "message": f"invalid status '{status}' "
                                              f"(valid: {STATUSES})"}
    from psycopg2.extras import RealDictCursor
    with get_db() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                UPDATE supplier_orders
                SET status = %s,
                    note = COALESCE(%s, note),
                    supplier_doc_ref = COALESCE(%s, supplier_doc_ref),
                    confirmed_at = CASE WHEN %s = 'confirmed' THEN NOW()
                                        ELSE confirmed_at END,
                    updated_at = NOW()
                WHERE id = %s
                RETURNING *
            """, (status, note, supplier_doc_ref, status, row_id))
            row = cur.fetchone()
            if not row:
                return {"status": "error", "message": f"supplier_order {row_id} not found"}
            cur.execute("""
                INSERT INTO order_events (order_id, event_type, event_data, source)
                VALUES (%s, 'supplier_order_status', %s, 'supplier_orders')
            """, (row["order_id"], json.dumps({
                "supplier_order_id": row_id, "warehouse": row["warehouse"],
                "status": status, "note": note, "doc_ref": supplier_doc_ref})))
            conn.commit()
            return {"status": "ok", "row": row}


def digest() -> Dict:
    """The 'what needs me today' view: counts by status + stale sent rows."""
    from psycopg2.extras import RealDictCursor
    with get_db() as conn:
        ensure_supplier_orders_table(conn)
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""SELECT status, COUNT(*) AS n FROM supplier_orders
                           GROUP BY status ORDER BY n DESC""")
            counts = {r["status"]: r["n"] for r in cur.fetchall()}
            cur.execute("""
                SELECT order_id, warehouse, status, sent_at, note
                FROM supplier_orders
                WHERE status IN ('sent', 'prepared')
                  AND sent_at < NOW() - interval '24 hours'
                ORDER BY sent_at ASC LIMIT 50
            """)
            stale = cur.fetchall()
            cur.execute("""
                SELECT order_id, warehouse, status, note
                FROM supplier_orders
                WHERE status IN ('blocked', 'discrepancy')
                ORDER BY updated_at DESC LIMIT 50
            """)
            needs_human = cur.fetchall()
    return {"status": "ok", "counts": counts,
            "unconfirmed_over_24h": stale, "needs_human": needs_human}
