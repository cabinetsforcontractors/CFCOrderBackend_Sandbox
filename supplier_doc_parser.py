"""
supplier_doc_parser.py
Per-supplier grammars that parse supplier sales-order / estimate documents back
into WEBSITE SKUs, plus the two-sided diff that verifies what the supplier
confirmed against what we sent (auto-ordering lane, William rulings 2026-07-16).

Decoded formats (SUPPLIER_ORDER_CHANNELS_20260716.md is THE spec):
  GHI       — Sales Order PDF: truncated prose descriptions, NO SKU; columns
              Ordered/Shipped/BackOrder/Price/Amount with fused numbers.
              Validated on 3 real SOs (PO 5177/5244/5568): every item line
              resolves or is explicitly flagged (markers/samples -> human).
  DuraStone — the NetSuite SO EMAIL is the document (clean HTML table
              Quantity|Item|Rate|Amount, no PDF). NW- = Natural Wood = our NSN.
              REVISION EMAILS (same SO# re-sent) = their mistake signal:
              SO112268 went $3,405.50 -> $1,955.50 in 40 min (B21 12 -> 2).
  LM / C&S / LI — split cabinets into universal-box + line-front pairs; parsed
              and folded back to cabinet bodies, diffed in body space
              (line-code -> our-prefix maps still pending William).

Accuracy bar: 99.5% send + receive. Lines this module cannot resolve are
returned as unresolved — they need a human, never silently dropped.
"""

import io
import json
import re
from datetime import datetime, timezone

# =============================================================================
# COMMON HELPERS
# =============================================================================

GHI_PREFIXES = ("AKS", "APW", "GRSH", "NOR", "SNS", "SNW")
DS_PREFIXES = ("NSN", "CMEN", "NBDS", "SIV")


def norm(s):
    return re.sub(r"\s+", "", str(s or "")).upper()


def norm_alnum(s):
    return re.sub(r"[^A-Z0-9]", "", norm(s))


def _money(s):
    try:
        return float(str(s).replace(",", "").replace("$", "").strip())
    except (TypeError, ValueError):
        return None


def _token_candidates(token):
    """Dialect-tolerant variants of a supplier token: as-is, without a trailing
    parenthetical (OGM8(4.5) -> OGM8), and each side of shared items (BM8/DSM8)."""
    t = norm(token)
    cands = [t]
    no_paren = re.sub(r"\(.*?\)$", "", t)
    if no_paren and no_paren != t:
        cands.append(no_paren)
    if "/" in t:
        cands.extend(p for p in t.split("/") if p)
    return cands


def build_reverse_map(conn):
    """(website prefix, normalized supplier token) -> website_sku, from
    rta_products.supplier_sku (SOT map). 'loose' is an alphanumeric-only index
    so OGM8(4.5") matches OGM8(4.5); ambiguous loose keys are disabled."""
    exact, loose = {}, {}
    with conn.cursor() as cur:
        cur.execute(
            """SELECT product_sku, supplier_sku FROM rta_products
               WHERE supplier_sku IS NOT NULL AND supplier_sku <> ''"""
        )
        rows = cur.fetchall()
    for product_sku, supplier_sku in rows:
        pre = (product_sku or "").split("-")[0].upper()
        exact.setdefault((pre, norm(supplier_sku)), product_sku)
        lk = (pre, norm_alnum(supplier_sku))
        if lk in loose and loose[lk] != product_sku:
            loose[lk] = None  # collision -> unusable
        else:
            loose[lk] = product_sku
    return {"exact": exact, "loose": loose}


def build_forward_map(conn):
    """website_sku -> supplier token (for outbound sheet/CSV generation)."""
    with conn.cursor() as cur:
        cur.execute(
            """SELECT product_sku, supplier_sku FROM rta_products
               WHERE supplier_sku IS NOT NULL AND supplier_sku <> ''"""
        )
        return {r[0]: r[1] for r in cur.fetchall()}


def rev_lookup(rev, prefix, token):
    """Resolve a supplier token to our website SKU for a given line prefix."""
    if not prefix or not token:
        return None
    for cand in _token_candidates(token):
        hit = rev["exact"].get((prefix, norm(cand)))
        if hit:
            return hit
    for cand in _token_candidates(token):
        hit = rev["loose"].get((prefix, norm_alnum(cand)))
        if hit:
            return hit
    return None


# =============================================================================
# GHI — Sales Order PDF grammar (prose descriptions, no SKUs)
# =============================================================================

# GHI color words -> our line prefix (descriptions truncate: "Stone Harbo")
GHI_COLOR = {
    "STONE HARBOR": "GRSH", "SONONA SAND": "SNS", "SONONA WHEAT": "SNW",
    "NANTUCKET": "NOR", "RUSTIC WALNUT": "APW", "RWS": "APW",
    "FRONTIER": "AKS", "FTS": "AKS", "SHG": "GRSH", "SNS": "SNS",
    "SNW": "SNW", "NTL": "NOR",
}

# GHI order-sheet tab per our line prefix (AKS=FTS proven on order 5155)
GHI_LINE_TAB = {"AKS": "FTS", "APW": "RWS", "GRSH": "SHG",
                "NOR": "NTL", "SNS": "SNS", "SNW": "SNW"}

# 90/96-high pantry composites: one website SKU = UCB base + wall-top pair.
# 84-high pantries are direct rows (e.g. WP2484FTS) — do NOT expand.
GHI_COMPOSITES = {
    "WP1890": ["UCB1854", "W183624"], "WP1896": ["UCB1854", "W184224"],
    "WP2490": ["UCB2454", "W243624"], "WP2496": ["UCB2454", "W244224"],
}

# bodies whose SOT token is missing/odd — GHI dialect fixes (proven 5155)
GHI_DIALECT = {"VAN48D-6": "V4821D", "VAN42D-6": "V4221D"}

# item line = desc + two 2dp floats (amount/price, possibly fused with qty)
# + qty + shipped + backorder
_GHI_LINE = re.compile(
    r"\n(.{4,60}?)\s*(\d[\d,]*\.\d{2})\s*(\d[\d,]*\.\d{2})(\d{1,3})\s+(\d{1,3})\s+(\d{1,3})"
)
_GHI_SKIP = ("Net Order", "Discount", "Freight", "Sales Tax", "Order Total",
             "ENTERED", ":")


def ghi_desc_to_tokens(desc):
    """Truncated-description grammar -> candidate GHI SKU tokens."""
    d = desc.upper().replace("''", "").strip()
    m = re.search(r"(\d+(?:\.\d+)?)\s*X\s*(\d+(?:\.\d+)?)", d)
    w, h = (m.group(1), m.group(2)) if m else (None, None)
    if "SAMPLE DOOR" in d:
        return ["SAMPLE"]
    if "UTILITY BASE" in d and w:
        return [f"UCB{int(float(w)):02d}{int(float(h)):02d}"]
    if "TOE KICK" in d:
        return ["TK8"]
    if "QUARTER ROUND" in d:
        return ["QR8"]
    if "SCRIBE" in d:
        return ["SCM7", "SM8"]
    if "CROWN" in d:
        return ["CM8"]
    if "FILLER" in d and w:
        return [f"F{int(float(w))}{int(float(h))}"]
    if "VANITY 3 DRAWER" in d and w:
        return [f"VDB{int(float(w))}{int(float(h))}"]
    if "VANITY" in d and "DRAWER" in d and w:
        return [f"V{int(float(w))}{int(float(h))}D"]
    if "VANITY" in d and w:
        return [f"V{int(float(w))}{int(float(h))}"]
    if "WALL" in d and w:
        return [f"W{int(float(w)):02d}{int(float(h)):02d}"]
    if ("BASE" in d or "CABINET" in d) and w:
        return [f"B{int(float(w)):02d}"]
    return []


def ghi_line_prefix(desc):
    """Match the (possibly truncated) color phrase -> our line prefix."""
    d = desc.upper()
    for k, v in GHI_COLOR.items():
        if k in d:
            return v
        for n in range(len(k), 2, -1):
            if k[:n] in d and (k[:n] == k or d.endswith(k[:n])):
                return v
    return None


def parse_ghi_pdf(data):
    """GHI Sales Order PDF bytes -> {'supplier','po','so_number','lines':[...]}"""
    from pypdf import PdfReader

    txt = "\n".join(p.extract_text() or "" for p in PdfReader(io.BytesIO(data)).pages)
    # PO is fused into the terms row: "Net 30SHIP COMPLETE5568" / "Net 305177"
    po = (re.search(r"Net\s*30[A-Z ]*?(\d{4}[A-Z]?)\s*\n", txt)
          or re.search(r"\n(\d{4}[A-Z]?)\s+SHIP", txt))
    # SO number is the first value after the header label block
    so = (re.search(r"Customer\s*Number:\s*\n(\d{6,8})\n", txt)
          or re.search(r"Order\s*Number[:\s]+(\d{4,8})", txt, re.IGNORECASE))
    lines = []
    for m in _GHI_LINE.finditer(txt):
        desc = m.group(1).strip()
        if any(k in desc for k in _GHI_SKIP):
            continue
        n1, n2 = (float(m.group(i).replace(",", "")) for i in (2, 3))
        qty, shipped, backorder = int(m.group(4)), int(m.group(5)), int(m.group(6))
        # amount = price * qty decides which fused float is which
        if abs(n1 - n2 * qty) < 0.02:
            amount, price = n1, n2
        elif abs(n2 - n1 * qty) < 0.02:
            amount, price = n2, n1
        else:
            amount, price = n1, n2  # unresolved -> math_ok False flags it
        lines.append({
            "desc": desc, "qty": qty, "price": price, "amount": amount,
            "shipped": shipped, "backorder": backorder,
            "marker": desc.startswith("**"),
            "math_ok": abs(amount - price * qty) < 0.02,
        })
    return {"supplier": "GHI", "po": po.group(1) if po else None,
            "so_number": so.group(1).lstrip("0") if so else None, "lines": lines}


# composite COMPONENT tokens (valid on GHI docs even without their own website SKU)
_GHI_COMPONENT_TOKENS = {t for parts in GHI_COMPOSITES.values() for t in parts}


def resolve_ghi_lines(parsed, rev):
    """Attach website_sku (or None) to each parsed GHI line."""
    prev_prefix = None
    for ln in parsed["lines"]:
        if ln.get("marker"):
            # "**USE 1 PIECE UNIT**" = one-piece pantry variant; this PRICED row
            # is the composite's top piece — needs a human to pair it.
            ln["ghi_tokens"], ln["line_prefix"] = [], prev_prefix
            ln["website_sku"] = None
            ln["note"] = "one-piece variant marker line (composite top) — verify manually"
            continue
        tokens = ghi_desc_to_tokens(ln["desc"])
        prefix = ghi_line_prefix(ln["desc"])
        ln["ghi_tokens"], ln["line_prefix"] = tokens, prefix
        ln["website_sku"] = None
        for t in tokens:
            hit = rev_lookup(rev, prefix, t)
            if hit:
                ln["website_sku"] = hit
                break
        if not ln["website_sku"] and prefix:
            # composite components (UCB1854, W244224...) have no website SKU of
            # their own — synthesize prefix-token so they diff against the
            # expanded sent side (expand_composites).
            for t in tokens:
                if norm(t) in _GHI_COMPONENT_TOKENS:
                    ln["website_sku"] = f"{prefix}-{norm(t)}"
                    ln["synthetic"] = True
                    break
        if not ln["website_sku"] and "SAMPLE" in [norm(t) for t in tokens]:
            ln["note"] = "sample door line — order via Misc sheet, verify manually"
        if prefix:
            prev_prefix = prefix
    return parsed


# =============================================================================
# DURASTONE — NetSuite Sales Order EMAIL grammar (HTML, no PDF)
# =============================================================================

# their line code -> our website prefix. NW validated 1:1 on PO 5568.
# CMEN / NBDS / SIV line codes still TBD from future SO emails.
DS_LINE_TO_PREFIX = {"NW": "NSN"}

_DS_ITEM_ROW = re.compile(
    r'<td[^>]*colspan="3"[^>]*>\s*(\d+)\s*</td>\s*'
    r'<td[^>]*colspan="12"[^>]*><span[^>]*>([^<]+)</span>(?:<br\s*/?>\s*([^<]*))?</td>\s*'
    r'<td[^>]*colspan="4"[^>]*>\$?\s*([\d,]+\.\d{2})\s*</td>\s*'
    r'<td[^>]*colspan="4"[^>]*>\$?\s*([\d,]+\.\d{2})\s*</td>',
    re.IGNORECASE | re.DOTALL)


def parse_durastone_email(html):
    """NetSuite SO email HTML -> {'supplier','so_number','po','total','lines'}"""
    so = re.search(r"#\s*(SO\d{4,8})", html)
    date = re.search(r"(\d{2}/\d{2}/\d{4})", html)
    po = None
    m = re.search(r"PO\s*#\s*</th>.*?<tr>(.*?)</tr>", html, re.DOTALL | re.IGNORECASE)
    if m:
        tds = re.findall(r"<td[^>]*>(.*?)</td>", m.group(1), re.DOTALL)
        if len(tds) >= 2:
            pm = re.search(r"(\d{3,6}[A-Z]?)", tds[1])
            po = pm.group(1) if pm else None
    totals = re.findall(r"Total</td>\s*<td[^>]*>\s*\$?\s*([\d,]+\.\d{2})",
                        html, re.IGNORECASE)
    block_m = re.search(r"<!--\s*start items\s*-->(.*?)<!--\s*end items\s*-->",
                        html, re.DOTALL | re.IGNORECASE)
    block = block_m.group(1) if block_m else html
    lines = []
    for qty, item, desc, rate, amount in _DS_ITEM_ROW.findall(block):
        code = item.split(",")[0].strip()          # "NW-B12, Natural Wood" -> NW-B12
        color = item.split(",", 1)[1].strip() if "," in item else ""
        line_code, _, body = code.partition("-")
        q, r, a = int(qty), _money(rate), _money(amount)
        lines.append({
            "item": code, "line_code": line_code.strip().upper(),
            "body": body.strip(), "color": color, "desc": (desc or "").strip(),
            "qty": q, "price": r, "amount": a,
            "math_ok": (r is not None and a is not None and abs(a - r * q) < 0.02),
        })
    return {"supplier": "DuraStone",
            "so_number": so.group(1) if so else None,
            "po": po, "order_date": date.group(1) if date else None,
            "total": _money(totals[-1]) if totals else None,
            "lines": lines}


def resolve_durastone_lines(parsed, rev):
    """Attach website_sku (or None) to each parsed DuraStone line."""
    for ln in parsed["lines"]:
        prefix = DS_LINE_TO_PREFIX.get(ln["line_code"])
        ln["line_prefix"] = prefix
        ln["website_sku"] = rev_lookup(rev, prefix, ln["body"]) if prefix else None
        if not prefix:
            ln["note"] = f"unknown DuraStone line code {ln['line_code']} (TBD: CMEN/NBDS/SIV)"
    return parsed


# =============================================================================
# TWO-SIDED DIFF — the accuracy backbone
# =============================================================================

def expand_composites(sent_lines, composites):
    """Expand composite website SKUs (WP2496 -> UCB2454 + W244224) so the sent
    side is in the same units the supplier confirms. composites keys are BODY
    tokens; expansion keeps the line prefix (AKS-WP2496 -> AKS-UCB2454...)."""
    if not composites:
        return list(sent_lines)
    out = []
    for ln in sent_lines:
        sku = ln.get("website_sku") or ""
        pre, _, body = sku.partition("-")
        parts = composites.get(norm(body))
        if parts:
            for p in parts:
                out.append({"website_sku": f"{pre}-{p}", "quantity": ln.get("quantity") or 1,
                            "composite_of": sku})
        else:
            out.append(dict(ln))
    return out


def two_sided_diff(sent_lines, supplier_lines):
    """Fingerprint both sides by website SKU and diff.

    sent_lines:     [{'website_sku', 'quantity'}]  — what we ordered
    supplier_lines: resolved parser lines — what the supplier confirmed
    Returns the discrepancy report; report['ok'] is True only when both sides
    match exactly AND every supplier line resolved AND no availability flags.
    """
    sent_q, recv_q, unresolved, flags = {}, {}, [], []
    for ln in sent_lines:
        sku = norm(ln.get("website_sku"))
        if sku:
            sent_q[sku] = sent_q.get(sku, 0) + int(ln.get("quantity") or 1)
    for ln in supplier_lines:
        sku = norm(ln.get("website_sku"))
        if not sku:
            unresolved.append({k: ln.get(k) for k in
                               ("item", "desc", "qty", "ghi_tokens", "line_prefix", "note")
                               if ln.get(k) is not None})
            continue
        recv_q[sku] = recv_q.get(sku, 0) + int(ln.get("qty") or 1)
        if ln.get("backorder"):
            flags.append(f"BACKORDER {ln['backorder']} x {sku} ({ln.get('desc', '')})")
        if ln.get("math_ok") is False:
            flags.append(f"PRICE MATH {sku}: qty {ln.get('qty')} x {ln.get('price')} != {ln.get('amount')}")
        if ln.get("memo"):
            flags.append(f"MEMO {sku}: {ln['memo']}")
        if ln.get("substituted_for"):
            flags.append(f"SUBSTITUTION {ln['substituted_for']} -> {sku}")
    matched, qty_mismatch = [], []
    missing = []      # we sent it, supplier did not confirm it
    unexpected = []   # supplier confirmed something we did not send
    for sku in sorted(set(sent_q) | set(recv_q)):
        s, r = sent_q.get(sku), recv_q.get(sku)
        if s and r:
            if s == r:
                matched.append({"sku": sku, "qty": s})
            else:
                qty_mismatch.append({"sku": sku, "sent_qty": s, "supplier_qty": r})
        elif s:
            missing.append({"sku": sku, "sent_qty": s})
        else:
            unexpected.append({"sku": sku, "supplier_qty": r})
    # pair up probable SKU-dialect drift (NSN-QR <-> NSN-QR8) so the human
    # reviewer sees candidates side by side — NEVER auto-accepted.
    substitutions = []
    for ms in missing:
        mp, _, mb = ms["sku"].partition("-")
        mb_a = norm_alnum(mb)
        for ux in unexpected:
            up, _, ub = ux["sku"].partition("-")
            ub_a = norm_alnum(ub)
            if mp == up and mb_a and ub_a and (
                    mb_a.startswith(ub_a) or ub_a.startswith(mb_a)):
                substitutions.append({"sent_sku": ms["sku"], "supplier_sku": ux["sku"],
                                      "sent_qty": ms["sent_qty"],
                                      "supplier_qty": ux["supplier_qty"]})
    return {
        "ok": not (qty_mismatch or missing or unexpected or unresolved or flags),
        "matched": matched, "qty_mismatch": qty_mismatch,
        "missing_at_supplier": missing, "unexpected_from_supplier": unexpected,
        "unresolved_supplier_lines": unresolved, "flags": flags,
        "possible_substitutions": substitutions,
        "sent_line_total": sum(sent_q.values()),
        "supplier_line_total": sum(recv_q.values()),
    }


def diff_ds_revisions(prev_lines, new_lines):
    """Diff two revisions of the same DuraStone SO by THEIR item code.
    Drops (missing items / qty decreases) are the mistake signal."""
    pq, nq = {}, {}
    for ln in prev_lines:
        pq[ln["item"]] = pq.get(ln["item"], 0) + int(ln.get("qty") or 0)
    for ln in new_lines:
        nq[ln["item"]] = nq.get(ln["item"], 0) + int(ln.get("qty") or 0)
    dropped = [{"item": i, "prev_qty": q} for i, q in pq.items() if i not in nq]
    added = [{"item": i, "new_qty": q} for i, q in nq.items() if i not in pq]
    qty_drops = [{"item": i, "prev_qty": pq[i], "new_qty": nq[i]}
                 for i in pq if i in nq and nq[i] < pq[i]]
    qty_increases = [{"item": i, "prev_qty": pq[i], "new_qty": nq[i]}
                     for i in pq if i in nq and nq[i] > pq[i]]
    return {"dropped": dropped, "qty_drops": qty_drops,
            "added": added, "qty_increases": qty_increases,
            "has_drops": bool(dropped or qty_drops)}


# =============================================================================
# GHI ORDER SHEET GENERATION (outbound) — proven on real order 5155
# =============================================================================

def make_ghi_sheets(items, template_bytes, po_number, fwd_map,
                    company="Cabinets For Contractors", ship_to=""):
    """Fill the GHI xlsx order sheet (Downloads 5707.xlsx format) from website-
    SKU items. items: [{'website_sku','quantity'}], GHI prefixes only.
    Returns (xlsx_bytes, report). Unplaced lines need Misc sheet / email note."""
    import openpyxl

    by_tab = {}
    unmapped = []
    for it in items:
        sku = norm(it.get("website_sku"))
        pre = sku.split("-")[0]
        tab = GHI_LINE_TAB.get(pre)
        if not tab:
            unmapped.append(sku)
            continue
        tok = fwd_map.get(it.get("website_sku")) or fwd_map.get(sku)
        if not tok:
            body = sku.split("-", 1)[1] if "-" in sku else sku
            tok = GHI_DIALECT.get(body, body)
        qty = int(it.get("quantity") or 1)
        if norm(tok) in GHI_COMPOSITES:
            for part in GHI_COMPOSITES[norm(tok)]:
                by_tab.setdefault(tab, []).append((part, qty, f"{sku} (composite {tok})"))
        else:
            by_tab.setdefault(tab, []).append((norm(tok), qty, sku))

    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    placed, unplaced = [], []
    for tab, wants in by_tab.items():
        if tab not in wb.sheetnames:
            unplaced.extend({"token": t, "qty": q, "source": s, "reason": f"no tab {tab}"}
                            for t, q, s in wants)
            continue
        ws = wb[tab]
        cells = {}  # token (tab suffix stripped) -> (row, order-qty col)
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200)):
            for c in row:
                if not isinstance(c.value, str):
                    continue
                m = re.match(rf"^([A-Z0-9/.\-]+?){tab}(\(FH\))?(\*.*)?$", norm(c.value))
                if m and c.column > 1:
                    cells.setdefault(m.group(1), (c.row, c.column - 1))
        for tok, qty, src in wants:
            hit = cells.get(tok)
            if not hit:
                for k in cells:
                    if k.startswith(tok) or tok.startswith(k):
                        hit = cells[k]
                        break
            if hit:
                r, col = hit
                cur = ws.cell(row=r, column=col).value
                ws.cell(row=r, column=col).value = (
                    (cur or 0) + qty if isinstance(cur, (int, float)) else qty)
                placed.append({"tab": tab, "token": tok, "qty": qty, "source": src})
            else:
                unplaced.append({"token": tok, "qty": qty, "source": src,
                                 "reason": "not on tab"})
        # header block: Date / Company / PO# / Ship To
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200)):
            for c in row:
                if not isinstance(c.value, str):
                    continue
                v = c.value.strip().rstrip(":").upper()
                tgt = ws.cell(row=c.row, column=c.column + 1)
                if v == "DATE":
                    tgt.value = datetime.now(timezone.utc).strftime("%m/%d/%Y")
                elif v == "COMPANY":
                    tgt.value = company
                elif v in ("PO#", "PO"):
                    tgt.value = str(po_number)
                elif v == "SHIP TO" and ship_to:
                    tgt.value = ship_to
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), {"placed": placed, "unplaced": unplaced,
                            "unmapped_prefix": unmapped,
                            "tabs": sorted(by_tab.keys())}


# =============================================================================
# DURASTONE REVISION TRIPWIRE (gmail path)
# =============================================================================

DS_SENDER = "system@sent-via.netsuite.com"


def ensure_supplier_so_table(conn):
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS supplier_sales_orders (
                id SERIAL PRIMARY KEY,
                supplier VARCHAR(50) NOT NULL,
                so_number VARCHAR(50) NOT NULL,
                po_number VARCHAR(50),
                email_message_id VARCHAR(120) UNIQUE,
                revision_num INTEGER DEFAULT 1,
                total DECIMAL(12,2),
                parsed_json TEXT,
                received_at TIMESTAMP WITH TIME ZONE,
                created_at TIMESTAMP WITH TIME ZONE DEFAULT NOW()
            )
        """)
        conn.commit()


def _get_email_html(message_id):
    """Fetch an email's text/html body (recursive part walk). Returns (headers,
    html) or (None, None). Uses gmail_sync's authenticated request helper."""
    import base64
    from gmail_sync import gmail_api_request

    data = gmail_api_request(f"messages/{message_id}", {"format": "full"})
    if not data:
        return None, None
    headers = {h["name"].lower(): h["value"]
               for h in data.get("payload", {}).get("headers", [])}

    def walk(part):
        if part.get("mimeType") == "text/html" and part.get("body", {}).get("data"):
            return base64.urlsafe_b64decode(part["body"]["data"]).decode("utf-8", errors="ignore")
        for p in part.get("parts", []) or []:
            found = walk(p)
            if found:
                return found
        return None

    return headers, walk(data.get("payload", {}))


def record_and_check_ds_email(conn, message_id, html, received_at=None):
    """Store a DuraStone NetSuite SO email; if the SO# was seen before this is
    a REVISION -> diff vs the prior revision and alert on drops.
    Returns {'status', 'so_number', 'revision_num', 'revision_diff'|None}."""
    parsed = parse_durastone_email(html)
    so_number = parsed.get("so_number")
    if not so_number or not parsed["lines"]:
        return {"status": "not_a_ds_so", "so_number": so_number}
    ensure_supplier_so_table(conn)
    with conn.cursor() as cur:
        cur.execute("SELECT 1 FROM supplier_sales_orders WHERE email_message_id = %s",
                    (message_id,))
        if cur.fetchone():
            return {"status": "already_recorded", "so_number": so_number}
        cur.execute("""SELECT parsed_json, revision_num FROM supplier_sales_orders
                       WHERE supplier = 'DuraStone' AND so_number = %s
                       ORDER BY revision_num DESC LIMIT 1""", (so_number,))
        prior = cur.fetchone()
        revision_num = (prior[1] + 1) if prior else 1
        cur.execute("""INSERT INTO supplier_sales_orders
                       (supplier, so_number, po_number, email_message_id,
                        revision_num, total, parsed_json, received_at)
                       VALUES ('DuraStone', %s, %s, %s, %s, %s, %s, %s)""",
                    (so_number, parsed.get("po"), message_id, revision_num,
                     parsed.get("total"), json.dumps(parsed),
                     received_at or datetime.now(timezone.utc)))
        conn.commit()
    result = {"status": "recorded", "so_number": so_number,
              "po_number": parsed.get("po"), "revision_num": revision_num,
              "total": parsed.get("total"), "line_count": len(parsed["lines"]),
              "revision_diff": None}
    if prior:
        prev = json.loads(prior[0])
        rdiff = diff_ds_revisions(prev.get("lines", []), parsed["lines"])
        rdiff["prev_total"], rdiff["new_total"] = prev.get("total"), parsed.get("total")
        result["revision_diff"] = rdiff
        try:
            with conn.cursor() as cur:
                cur.execute("""INSERT INTO order_events
                               (order_id, event_type, event_data, source)
                               VALUES (%s, 'ds_revision_detected', %s, 'gmail_sync')""",
                            (parsed.get("po") or so_number, json.dumps({
                                "so_number": so_number,
                                "revision_num": revision_num,
                                "diff": rdiff,
                            })))
                conn.commit()
        except Exception as e:
            conn.rollback()
            result["event_error"] = str(e)
    return result


def scan_durastone_emails(conn, hours_back=48):
    """Gmail scan for DuraStone NetSuite SO emails -> record + revision check.
    Wired as a guarded section of run_gmail_sync and as a manual endpoint."""
    from gmail_sync import gmail_configured, search_emails

    if not gmail_configured():
        return {"status": "skipped", "reason": "gmail_not_configured",
                "processed": 0, "alerts": 0, "errors": []}
    out = {"status": "ok", "processed": 0, "revisions": 0, "alerts": 0,
           "sos": [], "errors": []}
    try:
        messages = search_emails(
            f'newer_than:{int(hours_back)}h from:{DS_SENDER} subject:"Sales Order"')
    except Exception as e:
        return {"status": "error", "processed": 0, "alerts": 0,
                "errors": [f"search: {e}"]}
    for msg in messages:
        try:
            headers, html = _get_email_html(msg["id"])
            if not html:
                continue
            received_at = None
            try:
                from email.utils import parsedate_to_datetime
                if headers and headers.get("date"):
                    received_at = parsedate_to_datetime(headers["date"])
            except Exception:
                received_at = None
            res = record_and_check_ds_email(conn, msg["id"], html, received_at)
            if res.get("status") != "recorded":
                continue
            out["processed"] += 1
            out["sos"].append({k: res.get(k) for k in
                               ("so_number", "po_number", "revision_num", "total")})
            if res.get("revision_diff") is not None:
                out["revisions"] += 1
                if res["revision_diff"].get("has_drops"):
                    out["alerts"] += 1
        except Exception as e:
            out["errors"].append(f"{msg.get('id')}: {e}")
    return out


# =============================================================================
# BODY-SPACE VERIFICATION (LM / C&S / LI)
# LI, LM, C&S split cabinets into universal-box + line-front combos, and their
# line-code -> our-prefix maps are still pending William. Verification folds
# their components back to BODY tokens (B18, W2430...) and diffs those against
# the sent order's bodies + forward-map dialect tokens. Exact website-SKU
# resolution stays flagged until the line codes are ruled.
# =============================================================================

def _body_candidates(sku, supplier_sku=None):
    """Comparable body tokens for a sent order line: website body + the SOT
    forward-map dialect token (e.g. GSP-30-VANITYCOMBO-DL -> VSD3021L)."""
    out = set()
    sku = norm(sku)
    if "-" in sku:
        out.add(norm_alnum(sku.split("-", 1)[1]))
    out.add(norm_alnum(sku))
    if supplier_sku:
        for c in _token_candidates(supplier_sku):
            out.add(norm_alnum(c))
    return {c for c in out if c}


def body_space_diff(sent_lines, folded):
    """Diff folded supplier bodies vs sent lines in body space.

    sent_lines: [{'website_sku','quantity','supplier_sku'(optional)}]
    folded:     [{'body' or 'bodies' (shared alternates), 'qty', flags...}]
    """
    flags, unresolved = [], []
    sent = []  # (candidates, qty, sku)
    for ln in sent_lines:
        cands = _body_candidates(ln.get("website_sku"), ln.get("supplier_sku"))
        sent.append({"cands": cands, "qty": int(ln.get("quantity") or 1),
                     "sku": ln.get("website_sku"), "matched_qty": 0})
    recv = []
    for f in folded:
        bodies = f.get("bodies") or ([f["body"]] if f.get("body") else [])
        bodies = {norm_alnum(b) for b in bodies if b}
        if not bodies:
            unresolved.append(f)
            continue
        recv.append({"bodies": bodies, "qty": int(f.get("qty") or 0),
                     "raw": f.get("raw") or f.get("body"), "matched_qty": 0})
        for fl in f.get("flags") or []:
            flags.append(f"{'/'.join(sorted(bodies))}: {fl}")
    # greedy body matching (shared alternates match any member)
    for r in recv:
        need = r["qty"]
        for s in sent:
            if need <= 0:
                break
            if s["cands"] & r["bodies"]:
                take = min(need, s["qty"] - s["matched_qty"])
                if take > 0:
                    s["matched_qty"] += take
                    r["matched_qty"] += take
                    need -= take
    missing = [{"sku": s["sku"], "sent_qty": s["qty"],
                "unconfirmed_qty": s["qty"] - s["matched_qty"]}
               for s in sent if s["matched_qty"] < s["qty"]]
    unexpected = [{"body": "/".join(sorted(r["bodies"])), "supplier_qty": r["qty"],
                   "unmatched_qty": r["qty"] - r["matched_qty"]}
                  for r in recv if r["matched_qty"] < r["qty"]]
    return {
        "ok": not (missing or unexpected or unresolved or flags),
        "mode": "body_space",
        "matched_qty": sum(s["matched_qty"] for s in sent),
        "missing_at_supplier": missing,
        "unexpected_from_supplier": unexpected,
        "unresolved_supplier_lines": unresolved,
        "flags": flags,
        "sent_line_total": sum(s["qty"] for s in sent),
        "supplier_line_total": sum(r["qty"] for r in recv),
    }


# =============================================================================
# LOVE-MILESTONE — NetSuite QUOFL quote PDF (UBX- box + SB-DS- door split)
# =============================================================================

_LM_MEMOS = ("VERY LOW STOCK", "DROP SHIP", "LOW STOCK", "OUT OF STOCK")
_LM_TOKEN_PIECE = re.compile(r"^[A-Z0-9/().'\"-]+$")
_LM_FEES = ("PALLETS-L", "PALLETS-S", "SHIPPING")


def parse_lm_quote_pdf(data):
    """Milestone QUOFL#####.pdf -> lines. Item tokens wrap across lines
    ("SB- DS-3DB12/3VD\nB12"); pieces are concatenated until Ea/qty appears."""
    from pypdf import PdfReader

    txt = "\n".join(p.extract_text() or "" for p in PdfReader(io.BytesIO(data)).pages)
    quote = re.search(r"Document\s*#\s*:\s*(QUO[A-Z]*\d+)", txt)
    po = re.search(r"PO\s*:\s*(?:PO\s*)?(\d{3,6}[A-Z]?)", txt)
    lines = []
    for m in re.finditer(
            r"\n((?:UBX|SB|Pallets|Shipping)[\s\S]*?)\$([\d,]+\.\d{2})\s*\$([\d,]+\.\d{2})",
            txt):
        chunk = m.group(1)
        price, amount = _money(m.group(2)), _money(m.group(3))
        pieces = chunk.split()
        token_parts, qty, i = [], None, 0
        while i < len(pieces):
            p = pieces[i]
            if p == "Ea":
                i += 1
                continue
            if re.fullmatch(r"\d{1,4}", p):
                qty = int(p)
                break
            if token_parts and not _LM_TOKEN_PIECE.match(p):
                break
            token_parts.append(p)
            i += 1
        token = "".join(token_parts)
        if not token or qty is None:
            continue
        memo = [k for k in _LM_MEMOS if k in chunk.upper()]
        is_fee = any(token.upper().startswith(f) for f in _LM_FEES)
        lines.append({
            "item": token, "qty": qty, "price": price, "amount": amount,
            "memo": " / ".join(dict.fromkeys(memo)) if memo else "",
            "is_fee": is_fee,
            "math_ok": (price is not None and amount is not None
                        and abs(amount - price * qty) < 0.02),
        })
    return {"supplier": "Love-Milestone",
            "quote_number": quote.group(1) if quote else None,
            "po": po.group(1) if po else None, "lines": lines}


def fold_lm_lines(lines):
    """Fold UBX-{body} boxes + SB-DS-{body} door sets into cabinet bodies;
    shared door sets ("SB-DS-SB30/VS30") carry alternate bodies. Flags any
    box/door count imbalance — William: verify BOTH components per cabinet."""
    boxes, doors, folded = {}, {}, []
    for ln in lines:
        if ln.get("is_fee"):
            folded.append({"body": ln["item"], "qty": ln["qty"], "raw": ln["item"],
                           "flags": ["FEE LINE"], "fee": True})
            continue
        t = norm(ln["item"])
        flags = []
        if ln.get("memo"):
            flags.append(ln["memo"])
        if ln.get("math_ok") is False:
            flags.append(f"PRICE MATH {ln['item']}")
        if t.startswith("UBX-"):
            body = t[4:]
            boxes[body] = boxes.get(body, 0) + ln["qty"]
            if flags:
                folded.append({"bodies": body.split("/"), "qty": 0,
                               "raw": ln["item"], "flags": flags})
        elif t.startswith("SB-DS-") or re.match(r"^[A-Z]{2,4}-DS-", t):
            body = t.split("-DS-", 1)[1]
            doors[body] = doors.get(body, 0) + ln["qty"]
            if flags:
                folded.append({"bodies": body.split("/"), "qty": 0,
                               "raw": ln["item"], "flags": flags})
        else:
            # standalone accessory/panel (SB-ACM8, SB-FPV4296...) = complete item
            body = t.split("-", 1)[1] if "-" in t else t
            folded.append({"bodies": body.split("/"), "qty": ln["qty"],
                           "raw": ln["item"], "flags": flags})
    # reconcile boxes vs door sets — BOTH sides can carry shared alternates
    # (UBX-SB36/FSB36 box satisfies SB-DS-SB36/VS36 via the SB36 member)
    box_pool = [{"members": set(b.split("/")), "left": q, "key": b}
                for b, q in boxes.items()]
    boxes_only = bool(boxes) and not doors
    for dbody, dqty in doors.items():
        members = set(dbody.split("/"))
        pool = [bx for bx in box_pool if bx["members"] & members]
        available = sum(bx["left"] for bx in pool)
        use = min(available, dqty)
        folded.append({"bodies": sorted(members), "qty": dqty,
                       "raw": f"UBX+DS {dbody}",
                       "flags": ([] if available == dqty else
                                 [f"BOX/DOOR MISMATCH {dbody}: {available} boxes vs {dqty} door sets"])})
        for bx in pool:
            take = min(bx["left"], use)
            bx["left"] -= take
            use -= take
    for bx in box_pool:
        if bx["left"] <= 0:
            continue
        flags = [] if boxes_only else [f"BOX WITHOUT DOOR SET x{bx['left']}"]
        folded.append({"bodies": sorted(bx["members"]), "qty": bx["left"],
                       "raw": f"UBX-{bx['key']}", "flags": flags})
    if boxes_only:
        folded.append({"bodies": ["QUOTE-NOTE"], "qty": 0, "raw": "quote summary",
                       "flags": ["QUOTE HAS BOXES ONLY — no door sets on this quote"]})
    return folded


# =============================================================================
# CABINET & STONE — Odoo Quotation PDF (Combo groups + bracket substitutions)
# =============================================================================

_CS_COMBO = re.compile(r"\n([A-Z0-9/.\-]+)-Combo\s*\n")
_CS_COMPONENT = re.compile(
    r"\n(Houston|Dallas|CA)\s*\n(\d+\.\d{2})\s*\n(.+?)\n([\s\S]{0,220}?)"
    r"([\d,]+\.\d{2})\s*\n(\d{1,2}\.\d{2})\s*\n([\d,.]+)")


def parse_cs_quotation_pdf(data):
    """C&S-Houston Quotation (Ref S#####) PDF -> component lines with combo
    membership and bracketed substitutions ("[requested] supplied")."""
    from pypdf import PdfReader

    txt = "\n".join(p.extract_text() or "" for p in PdfReader(io.BytesIO(data)).pages)
    ref = re.search(r"Quotation\s*#\s*\n?\s*(S\d{4,8})", txt)
    po = re.search(r"PO\s*:\s*\n?\s*(\d{3,6}[A-Z]?)", txt)
    combos = [(m.start(), m.group(1)) for m in _CS_COMBO.finditer(txt)]
    lines = []
    for m in _CS_COMPONENT.finditer(txt):
        raw_item = m.group(3).strip()
        requested = None
        bm = re.match(r"\[([^\]]+)\]\s*(.+)$", raw_item)
        if bm:
            requested, raw_item = bm.group(1).strip(), bm.group(2).strip()
        combo = None
        for pos, name in combos:
            if pos < m.start():
                combo = name
            else:
                break
        # only components that belong to the latest combo header keep it
        if combo and not (norm_alnum(combo.split("-")[-1]) in norm_alnum(raw_item)
                          or raw_item.upper().startswith("REGULAR")):
            combo = None
        lines.append({
            "item": raw_item, "qty": int(float(m.group(2))),
            "desc": m.group(4).strip()[:120],
            "price": _money(m.group(5)), "disc": _money(m.group(6)),
            "amount": _money(m.group(7)),
            "combo": combo, "substituted_for": requested,
        })
    return {"supplier": "Cabinet & Stone",
            "quote_number": ref.group(1) if ref else None,
            "po": po.group(1) if po else None, "lines": lines}


def fold_cs_lines(lines):
    """Fold Combo groups (NB-{b}-Box + [Regular w/Shelf] + {line}-{b}-Door) to
    one body per cabinet; verify BOTH components present (William 2026-07-16).
    Standalone items pass through as their own body."""
    combos, folded = {}, []
    for ln in lines:
        t = norm(ln["item"])
        flags = []
        if ln.get("substituted_for"):
            flags.append(f"SUBSTITUTION {ln['substituted_for']} -> {ln['item']}")
        m = re.match(r"^([A-Z0-9]+)-([A-Z0-9/.]+)-(BOX|DOOR)$", t)
        if m:
            body = m.group(2)
            slot = m.group(3)
            c = combos.setdefault(body, {"BOX": 0, "DOOR": 0, "flags": []})
            c[slot] += ln["qty"]
            c["flags"].extend(flags)
        elif t.startswith("REGULAR"):
            continue  # $0 shelf-config line inside a combo
        else:
            body = t.split("-", 1)[1] if "-" in t else t
            folded.append({"bodies": body.split("/"), "qty": ln["qty"],
                           "raw": ln["item"], "flags": flags})
    for body, c in combos.items():
        flags = list(dict.fromkeys(c["flags"]))
        if c["BOX"] != c["DOOR"]:
            flags.append(f"BOX/DOOR MISMATCH {body}: {c['BOX']} boxes vs {c['DOOR']} doors")
        folded.append({"bodies": body.split("/"), "qty": max(c["BOX"], c["DOOR"]),
                       "raw": f"Combo {body}", "flags": flags})
    return folded


# =============================================================================
# LI (Cabinetry Distribution) — QuickBooks Estimate/Invoice PDF
# =============================================================================

_LI_LINE = re.compile(
    r"^(\d{1,4})\s+([A-Z0-9][A-Z0-9/().'\"-]*(?:-[A-Z0-9/().'\"]+)*)"
    r"(?:\s+(PC\d{3,6}))?\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})T?\s*$",
    re.MULTILINE)


def parse_li_estimate_pdf(data):
    """LI QuickBooks Estimate/Invoice PDF -> lines. ACTIVITY = their SKU
    (SJ-{body}-DF door/face + {body}-BOX); composite talls = UT#####-BOX/DF
    pairs with our PC-code in the DESCRIPTION column."""
    from pypdf import PdfReader

    txt = "\n".join(p.extract_text() or "" for p in PdfReader(io.BytesIO(data)).pages)
    est = re.search(r"(?:ESTIMATE|INVOICE)\s*#?\s*(\d{3,6})", txt)
    po = re.search(r"PO\s*\n?\s*(\d{3,6}[A-Z]?)", txt)
    total = re.search(r"TOTAL\s*\$([\d,]+\.\d{2})", txt)
    discount = re.search(r"DISCOUNT\s*(\d{1,2})%", txt)
    lines = []
    for m in _LI_LINE.finditer(txt):
        lines.append({
            "item": m.group(2), "qty": int(m.group(1)),
            "pc_code": m.group(3) or "",
            "price": _money(m.group(4)), "amount": _money(m.group(5)),
        })
    return {"supplier": "LI",
            "estimate_number": est.group(1) if est else None,
            "po": po.group(1) if po else None,
            "total": _money(total.group(1)) if total else None,
            "discount_pct": int(discount.group(1)) if discount else None,
            "lines": lines}


def fold_li_lines(lines):
    """Fold {line}-{body}-DF + {body}-BOX pairs into bodies; UT component
    pairs with a PC code fold into that PC body (our composite SKU)."""
    dfs, boxes, pc_groups, folded = {}, {}, {}, []
    for ln in lines:
        t = norm(ln["item"])
        if ln.get("pc_code"):
            g = pc_groups.setdefault(norm(ln["pc_code"]), {"qtys": [], "items": []})
            g["qtys"].append(ln["qty"])
            g["items"].append(ln["item"])
            continue
        if t.endswith("-BOX"):
            body = t[:-4]
            boxes[body] = boxes.get(body, 0) + ln["qty"]
        elif t.endswith("-DF"):
            body = t[:-3].split("-", 1)[1] if "-" in t[:-3] else t[:-3]
            dfs[body] = dfs.get(body, 0) + ln["qty"]
        else:
            body = t.split("-", 1)[1] if "-" in t else t
            folded.append({"bodies": [body], "qty": ln["qty"], "raw": ln["item"],
                           "flags": []})
    for body in sorted(set(dfs) | set(boxes)):
        d, b = dfs.get(body, 0), boxes.get(body, 0)
        flags = []
        if d != b:
            flags.append(f"DF/BOX MISMATCH {body}: {d} door-faces vs {b} boxes")
        folded.append({"bodies": [body], "qty": max(d, b),
                       "raw": f"DF+BOX {body}", "flags": flags})
    for pc, g in pc_groups.items():
        qty = min(g["qtys"]) if g["qtys"] else 0
        flags = []
        if len(set(g["qtys"])) > 1:
            flags.append(f"PC COMPONENT QTY MISMATCH {pc}: {g['items']} {g['qtys']}")
        folded.append({"bodies": [pc], "qty": qty, "raw": f"composite {pc}",
                       "flags": flags})
    return folded
